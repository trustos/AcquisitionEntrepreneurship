#!/usr/bin/env python3
"""
Save one acquisition evaluation so nothing is lost and deals stay comparable.

Writes a per-deal folder (scorecard.md + record.json with all sources) and
appends/updates one row in the combined deal log (deal-log.xlsx + deal-log.csv).

TEAM USE: the per-deal record.json files are the source of truth and never
conflict in git (one file per evaluation). The combined deal-log.* is a DERIVED
artifact — if it ever conflicts on a shared repo, don't hand-merge it, just run
`rebuild_log.py --dir deals/` to regenerate it from every record.json.

Usage:
    python scripts/save_evaluation.py --record record.json --dir /path/to/deals [--analyst "Yavor"]
    cat record.json | python scripts/save_evaluation.py --dir ./deals

Record keys (see SKILL.md Phase 8): target, date, analyst, type, verdict, score,
completeness_pct, one_line_reason, asking_price, dimensions, gates, key_unknowns,
seller_questions, inputs_provided, sources, scorecard_markdown.
Re-running for the same target+date(+analyst) updates that entry, not a duplicate.
"""
import argparse
import csv
import datetime
import json
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAVE_XLSX = True
except Exception:
    HAVE_XLSX = False

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import normalize_report  # canonical, import-ready report shape
import score as scorer   # deterministic re-check (S5)


def verify_against_scorer(rec):
    """S5: re-derive the score from the record's OWN dimensions+gates and return
    a list of mismatches — so a hand-edited or mis-transcribed score/verdict can't
    be saved silently. Only meaningful for v2 records (score.py is the v2 scorer)."""
    if rec.get("scoring_version") != "v2":
        return []
    deal = rec.get("deal", {}) or {}
    computed = scorer.evaluate({
        "target": rec.get("target"),
        "budget_ceiling": deal.get("budget_ceiling_usd"),
        "asking_price": deal.get("asking_price_usd"),
        "financing_possible": deal.get("financing_possible"),
        "dimensions": rec.get("dimensions", {}) or {},
        "gates": rec.get("gates", {}) or {},
    })
    out = []
    if rec.get("score") is not None and abs(float(rec["score"]) - computed["total"]) > 0.5:
        out.append(f"score {rec['score']} != recomputed {computed['total']}")
    rec_tier = (rec.get("verdict_detail") or {}).get("tier")
    if rec_tier and rec_tier != computed["tier"]:
        out.append(f"verdict tier {rec_tier} != recomputed {computed['tier']}")
    return out

COLUMNS = ["Date", "Target", "Analyst", "Type", "Verdict", "Score",
           "Completeness %", "Asking price", "One-line reason", "Folder"]
_WIDTHS = [12, 24, 12, 22, 22, 8, 14, 16, 46, 30]


def slugify(s):
    s = re.sub(r"^https?://", "", str(s or "target").strip().lower())
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "target"


def folder_for(rec):
    f = f"{slugify(rec.get('target'))}_{rec.get('date')}"
    if rec.get("analyst"):
        f += "_" + slugify(rec["analyst"])
    return f


def row_from(rec):
    return [rec.get("date", ""), rec.get("target", ""), rec.get("analyst", ""),
            rec.get("type", ""), rec.get("verdict", ""), rec.get("score", ""),
            rec.get("completeness_pct", ""), rec.get("asking_price") or "",
            rec.get("one_line_reason", ""), folder_for(rec)]


def write_csv(path, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        w.writerows(rows)


def write_xlsx(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "deals"
    ws.append(COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3B57")
    ws.freeze_panes = "A2"
    for i, wd in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    for row in rows:
        ws.append(row)
    wb.save(path)


def _read_csv_rows(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="") as f:
        rows = list(csv.reader(f))
    if rows and rows[0] == COLUMNS:
        rows = rows[1:]
    return rows


def upsert(base, row):
    """Add or replace one row (keyed on the unique Folder column) in both logs."""
    folder = row[-1]
    rows = [r for r in _read_csv_rows(os.path.join(base, "deal-log.csv"))
            if not (r and r[-1] == folder)]
    rows.append(row)
    write_csv(os.path.join(base, "deal-log.csv"), rows)
    if HAVE_XLSX:
        write_xlsx(os.path.join(base, "deal-log.xlsx"), rows)


def scorecard_md(rec):
    if rec.get("scorecard_markdown"):
        return rec["scorecard_markdown"]
    lines = [f"# Acquisition Scorecard — {rec.get('target', 'target')}", "",
             f"**Verdict:** {rec.get('verdict', '?')} — {rec.get('score', '?')}/100 "
             f"(completeness {rec.get('completeness_pct', '?')}%)"
             + (f"  ·  analyst: {rec['analyst']}" if rec.get("analyst") else ""), "",
             rec.get("one_line_reason", ""), ""]
    for k, v in (rec.get("dimensions") or {}).items():
        if lines[-1] != "## Scores":
            lines.append("## Scores")
        lines.append(f"- {k}: {v.get('score', '?')}/5 ({v.get('confidence', '?')}) {v.get('note', '')}".rstrip())
    if rec.get("dimensions"):
        lines.append("")
    if rec.get("key_unknowns"):
        lines.append("## Key unknowns")
        lines += [f"- {u}" for u in rec["key_unknowns"]] + [""]
    if rec.get("seller_questions"):
        lines.append("## Seller questions")
        lines += [f"{i + 1}. {q}" for i, q in enumerate(rec["seller_questions"])] + [""]
    if rec.get("sources"):
        lines.append("## Sources")
        for s in rec["sources"]:
            lines.append(f"- [{s.get('title', s.get('url', ''))}]({s.get('url', '')})" if isinstance(s, dict) else f"- {s}")
        lines.append("")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description="Save an acquisition evaluation and update the deal log.")
    ap.add_argument("--record", "-r", help="Path to record JSON (default: stdin)")
    ap.add_argument("--dir", "-d", default="deals", help="Deals base folder (default: ./deals)")
    ap.add_argument("--analyst", help="Who ran this evaluation (attribution); overrides record.analyst")
    ap.add_argument("--force", action="store_true", help="Save even if the stored score/verdict disagree with the scorer (S5)")
    args = ap.parse_args()

    if args.record:
        with open(args.record) as f:
            rec = json.load(f)
    else:
        raw = sys.stdin.read().strip()
        if not raw:
            sys.exit("No record provided (use --record FILE or pipe JSON via stdin).")
        rec = json.loads(raw)

    rec.setdefault("date", datetime.date.today().isoformat())
    if args.analyst:
        rec["analyst"] = args.analyst
    base = os.path.abspath(args.dir)
    folder = folder_for(rec)
    deal_dir = os.path.join(base, folder)
    os.makedirs(deal_dir, exist_ok=True)

    rec = normalize_report.to_canonical(rec, folder)   # make the report import-ready
    issues = normalize_report.validate(rec)
    if issues:
        print("  ! report validation: " + "; ".join(issues))

    mismatches = verify_against_scorer(rec)   # S5
    if mismatches and not args.force:
        print("\n  ✗ REFUSING TO SAVE — the record disagrees with the scorer:")
        for m in mismatches:
            print(f"      - {m}")
        print("    Re-run score.py on these dimensions/gates and fix the record, or pass --force.\n")
        sys.exit(2)
    if mismatches:
        print("  ! saved WITH --force despite scorer mismatch: " + "; ".join(mismatches))

    with open(os.path.join(deal_dir, "record.json"), "w") as f:
        json.dump(rec, f, indent=2, ensure_ascii=False)
    with open(os.path.join(deal_dir, "scorecard.md"), "w") as f:
        f.write(scorecard_md(rec))

    upsert(base, row_from(rec))
    print("Saved evaluation:")
    print(f"  - {os.path.join(deal_dir, 'scorecard.md')}")
    print(f"  - {os.path.join(deal_dir, 'record.json')}")
    print(f"  - {os.path.join(base, 'deal-log.csv')} (row added/updated)")
    print(f"  - {os.path.join(base, 'deal-log.xlsx')} (row added/updated)" if HAVE_XLSX
          else "  - deal-log.xlsx SKIPPED (pip install openpyxl --break-system-packages)")


if __name__ == "__main__":
    main()
